from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from authentication.models import *
from allotment.models import *
from .serializers import *
from authentication.permissions import *
from rest_framework.permissions import IsAuthenticated
from .models import *
from rest_framework_simplejwt.authentication import JWTAuthentication
from django.utils import timezone
from django.db.models import Count, Sum
from django.db.models import Q
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from io import BytesIO
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side



User = get_user_model()

class SeatMatrixView(APIView):
    permission_classes = [IsAuthenticated, IsStaffUser]

    def get(self, request):
        year = request.query_params.get('year')
        gender = request.query_params.get('gender')
        if not year or not gender:
            return Response({"error": "Year and gender are required"}, status=status.HTTP_400_BAD_REQUEST)
        try:
            seat_matrix = SeatMatrix.objects.get(year=year, gender=gender)
            serializer = SeatMatrixSerializer(seat_matrix)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except SeatMatrix.DoesNotExist:
            return Response({"error": "Seat matrix not found"}, status=status.HTTP_404_NOT_FOUND)

    def post(self, request):
        year = request.data.get('year')
        gender = request.data.get('gender')
        if year and gender:
            try:
                existing_matrix = SeatMatrix.objects.get(year=year, gender=gender)
                if existing_matrix.reserved_seats:
                    existing_matrix.reserved_seats.delete()
                existing_matrix.delete()
            except SeatMatrix.DoesNotExist:
                pass

        serializer = SeatMatrixSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class OpenRegistrationsView(APIView):
    permission_classes = [IsAuthenticated, IsManager]

    def get(self, request):
        year = request.query_params.get("year")
        if year:
            dates = SelectDates.objects.filter(
                event="Open Registrations", year=year.lower().replace(" ", "_")
            )
        else:
            dates = SelectDates.objects.filter(event="Open Registrations")
        serializer = SelectDatesSerializer(dates, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def post(self, request):
        # Delete completed events
        now = timezone.now()
        SelectDates.objects.filter(
            event="Open Registrations",
            end_date__lt=now,
            end_date__isnull=False
        ).delete()

        data = request.data.copy()
        data["event"] = "Open Registrations"
        years = data.get("years", [])

        if not years:
            return Response(
                {"error": "At least one year is required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Check for existing ongoing or upcoming events
        existing_events = SelectDates.objects.filter(
            event="Open Registrations",
            year__in=[year.lower().replace(" ", "_") for year in years]
        )
        for event in existing_events:
            is_ongoing = event.start_date <= now and (event.end_date is None or event.end_date >= now)
            is_upcoming = event.start_date > now
            if is_ongoing or is_upcoming:
                status_str = "ongoing" if is_ongoing else "upcoming"
                return Response(
                    {"error": f"Cannot create new event: {status_str} Open Registrations event exists for year {event.year}"},
                    status=status.HTTP_400_BAD_REQUEST
                )

        serializer = SelectDatesSerializer(data=data, context={"request": request})
        if not serializer.is_valid():
            print("Serializer errors:", serializer.errors)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        serializer.save()
        return Response(
            {"message": "Open Registrations dates saved successfully"},
            status=status.HTTP_200_OK
        )

    def delete(self, request):
        year = request.query_params.get("year")
        if not year:
            return Response(
                {"error": "Year is required to delete an event"},
                status=status.HTTP_400_BAD_REQUEST
            )
        try:
            instance = SelectDates.objects.get(
                event="Open Registrations", year=year.lower().replace(" ", "_")
            )
            instance.delete()
            return Response(
                {"message": "Open Registrations event deleted successfully"},
                status=status.HTTP_200_OK
            )
        except SelectDates.DoesNotExist:
            return Response(
                {"error": "No Open Registrations event found for the specified year"},
                status=status.HTTP_404_NOT_FOUND
            )

class OpenRoomPreferencesView(APIView):
    permission_classes = [IsAuthenticated, IsManager]

    def get(self, request):
        year = request.query_params.get("year")
        if year:
            dates = SelectDates.objects.filter(
                event="Open Room Preferences", year=year.lower().replace(" ", "_")
            )
        else:
            dates = SelectDates.objects.filter(event="Open Room Preferences")
        serializer = SelectDatesSerializer(dates, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def post(self, request):
        # Delete completed events
        now = timezone.now()
        SelectDates.objects.filter(
            event="Open Room Preferences",
            end_date__lt=now,
            end_date__isnull=False
        ).delete()

        data = request.data.copy()
        data["event"] = "Open Room Preferences"
        years = data.get("years", [])

        if not years:
            return Response(
                {"error": "At least one year is required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Check for existing ongoing or upcoming events
        existing_events = SelectDates.objects.filter(
            event="Open Room Preferences",
            year__in=[year.lower().replace(" ", "_") for year in years]
        )
        for event in existing_events:
            is_ongoing = event.start_date <= now and (event.end_date is None or event.end_date >= now)
            is_upcoming = event.start_date > now
            if is_ongoing or is_upcoming:
                status_str = "ongoing" if is_ongoing else "upcoming"
                return Response(
                    {"error": f"Cannot create new event: {status_str} Open Room Preferences event exists for year {event.year}"},
                    status=status.HTTP_400_BAD_REQUEST
                )

        serializer = SelectDatesSerializer(data=data, context={"request": request})
        if not serializer.is_valid():
            print("Serializer errors:", serializer.errors)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        serializer.save()
        return Response(
            {"message": "Open Room Preferences dates saved successfully"},
            status=status.HTTP_200_OK
        )

    def delete(self, request):
        year = request.query_params.get("year")
        if not year:
            return Response(
                {"error": "Year is required to delete an event"},
                status=status.HTTP_400_BAD_REQUEST
            )
        try:
            instance = SelectDates.objects.get(
                event="Open Room Preferences", year=year.lower().replace(" ", "_")
            )
            instance.delete()
            return Response(
                {"message": "Open Room Preferences event deleted successfully"},
                status=status.HTTP_200_OK
            )
        except SelectDates.DoesNotExist:
            return Response(
                {"error": "No Open Room Preferences event found for the specified year"},
                status=status.HTTP_404_NOT_FOUND
            )

class PendingStudentsView(APIView):
    permission_classes = [IsAuthenticated, IsManager]

    def get(self, request):
        pending_students = StudentDataEntry.objects.filter(verified=None)
        year = request.query_params.get('class_name')
        if year and year in dict(StudentDataEntry.CLASS_CHOICES).keys():
            pending_students = pending_students.filter(class_name=year)
        serializer = StudentDataEntrySerializer(pending_students, many=True)
        return Response({"data": serializer.data}, status=status.HTTP_200_OK)

    def post(self, request):
        roll_no = request.data.get('roll_no')
        verified = request.data.get('verified')
        if verified not in [True, False]:
            return Response(
                {"error": "Invalid value for 'verified' (must be True or False)."},
                status=status.HTTP_400_BAD_REQUEST
            )
        try:
            student = StudentDataEntry.objects.get(roll_no=roll_no, verified=None)
            student.verified = verified
            student.save()
            return Response(
                {"message": "Student status updated successfully."},
                status=status.HTTP_200_OK
            )
        except StudentDataEntry.DoesNotExist:
            return Response(
                {"error": "Student not found or already processed."},
                status=status.HTTP_404_NOT_FOUND
            )

class VerifiedStudentsView(APIView):
    permission_classes = [IsAuthenticated, IsManager]

    def get(self, request):
        verified_students = StudentDataEntry.objects.filter(verified=True)
        year = request.query_params.get('class_name')
        if year and year in dict(StudentDataEntry.CLASS_CHOICES).keys():
            verified_students = verified_students.filter(class_name=year)
        serializer = StudentDataEntrySerializer(verified_students, many=True)
        return Response({"data": serializer.data}, status=status.HTTP_200_OK)

    def post(self, request):
        roll_no = request.data.get('roll_no')
        try:
            student = StudentDataEntry.objects.get(roll_no=roll_no, verified=True)
            student.verified = False
            student.save()
            return Response(
                {"message": "Student rejected successfully."},
                status=status.HTTP_200_OK
            )
        except StudentDataEntry.DoesNotExist:
            return Response(
                {"error": "Student not found or not verified."},
                status=status.HTTP_404_NOT_FOUND
            )

class RejectedStudentsView(APIView):
    permission_classes = [IsAuthenticated, IsManager]

    def get(self, request):
        rejected_students = StudentDataEntry.objects.filter(verified=False)
        year = request.query_params.get('class_name')
        if year and year in dict(StudentDataEntry.CLASS_CHOICES).keys():
            rejected_students = rejected_students.filter(class_name=year)
        serializer = StudentDataEntrySerializer(rejected_students, many=True)
        return Response({"data": serializer.data}, status=status.HTTP_200_OK)

    def post(self, request):
        roll_no = request.data.get('roll_no')
        try:
            student = StudentDataEntry.objects.get(roll_no=roll_no, verified=False)
            student.verified = True
            student.save()
            return Response(
                {"message": "Student verified successfully."},
                status=status.HTTP_200_OK
            )
        except StudentDataEntry.DoesNotExist:
            return Response(
                {"error": "Student not found or not rejected."},
                status=status.HTTP_404_NOT_FOUND
            )

class StudentDetailView(APIView):
    permission_classes = [IsAuthenticated, IsManager]

    def get(self, request, roll_no):
        try:
            student = StudentDataEntry.objects.get(roll_no=roll_no)
            serializer = StudentDetailSerializer(student)
            return Response({"data": serializer.data}, status=status.HTTP_200_OK)
        except StudentDataEntry.DoesNotExist:
            return Response(
                {"error": "Student not found."},
                status=status.HTTP_404_NOT_FOUND
            )

class StudentsByYearView(APIView):
    permission_classes = [IsAuthenticated, IsManager]

    def get(self, request):
        year = request.query_params.get('class_name')
        if year not in dict(StudentDataEntry.CLASS_CHOICES).keys():
            return Response(
                {"error": "Invalid year."},
                status=status.HTTP_400_BAD_REQUEST
            )
        students = StudentDataEntry.objects.filter(class_name=year)
        serializer = StudentDataEntrySerializer(students, many=True)
        return Response({"data": serializer.data}, status=status.HTTP_200_OK)

class WardensView(APIView):
    permission_classes = [IsAuthenticated, IsRector]

    def get(self, request):
        wardens = User.objects.filter(user_type="warden")
        serializer = UserSerializer(wardens, many=True)
        return Response({"data": serializer.data}, status=status.HTTP_200_OK)

    def post(self, request):
        serializer = UserCreateSerializer(data={**request.data, "user_type": "warden"})
        if serializer.is_valid():
            user = serializer.save()
            return Response(
                {"message": "Warden created successfully.", "data": UserSerializer(user).data},
                status=status.HTTP_201_CREATED
            )
        return Response(
            {"error": serializer.errors},
            status=status.HTTP_400_BAD_REQUEST
        )

    def delete(self, request, user_id):
        try:
            user = User.objects.get(id=user_id, user_type="warden")
            user.delete()
            return Response(
                {"message": "Warden deleted successfully."},
                status=status.HTTP_200_OK
            )
        except User.DoesNotExist:
            return Response(
                {"error": "Warden not found."},
                status=status.HTTP_404_NOT_FOUND
            )

class ManagersView(APIView):
    permission_classes = [IsAuthenticated, IsRector]

    def get(self, request):
        managers = User.objects.filter(user_type="manager")
        serializer = UserSerializer(managers, many=True)
        return Response({"data": serializer.data}, status=status.HTTP_200_OK)

    def post(self, request):
        serializer = UserCreateSerializer(data={**request.data, "user_type": "manager"})
        if serializer.is_valid():
            user = serializer.save()
            return Response(
                {"message": "Manager created successfully.", "data": UserSerializer(user).data},
                status=status.HTTP_201_CREATED
            )
        return Response(
            {"error": serializer.errors},
            status=status.HTTP_400_BAD_REQUEST
        )

    def delete(self, request, user_id):
        try:
            user = User.objects.get(id=user_id, user_type="manager")
            user.delete()
            return Response(
                {"message": "Manager deleted successfully."},
                status=status.HTTP_200_OK
            )
        except User.DoesNotExist:
            return Response(
                {"error": "Manager not found."},
                status=status.HTTP_404_NOT_FOUND
            )

class GetBranchesView(APIView):
    permission_classes = [IsAuthenticated, IsManager]

    def get(self, request):
        try:
            year = request.query_params.get('year', 'fy')
            gender = request.query_params.get('gender', 'male')
            try:
                seat_matrix = SeatMatrix.objects.get(year=year, gender=gender)
                branches = list(seat_matrix.branch_seats.keys())
                return Response({"branches": branches}, status=status.HTTP_200_OK)
            except SeatMatrix.DoesNotExist:
                return Response({"branches": []}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response(
                {"error": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class AllotBranchRanksView(APIView):
    permission_classes = [IsAuthenticated, IsManager]

    def post(self, request):
        try:
            year = request.data.get('year')
            gender = request.data.get('gender', 'male')
            valid_years = [choice[0] for choice in StudentDataEntry.CLASS_CHOICES]
            if not year or year not in valid_years:
                return Response(
                    {"error": f"Invalid year. Must be one of: {', '.join(valid_years)}"},
                    status=status.HTTP_400_BAD_REQUEST
                )
            valid_genders = [choice[0] for choice in StudentDataEntry.GENDER_CHOICES]
            if gender not in valid_genders:
                return Response(
                    {"error": f"Invalid gender. Must be one of: {', '.join(valid_genders)}"},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Filter students, excluding manually modified ones
            students = StudentDataEntry.objects.filter(
                class_name=year,
                gender=gender,
                is_manually_modified=False
            ).select_related('branch')

            if not students.exists():
                return Response(
                    {"message": f"No {year} {gender} students found."},
                    status=status.HTTP_200_OK
                )

            # Group students by branch
            branch_groups = {}
            for student in students:
                branch_id = student.branch_id
                if branch_id not in branch_groups:
                    branch_groups[branch_id] = []
                branch_groups[branch_id].append(student)

            # Assign branch ranks
            total_updated = 0
            for branch_id, students in branch_groups.items():
                sorted_students = sorted(students, key=lambda s: (s.cgpa if s.cgpa is not None else 0), reverse=True)
                for rank, student in enumerate(sorted_students, start=1):
                    if student.branch_rank != rank:
                        student.branch_rank = rank
                        student.save(update_fields=['branch_rank'])
                        total_updated += 1

            # Call SelectStudentsView, passing the same request
            select_view = SelectStudentsView()
            select_response = select_view.post(request)

            if select_response.status_code != 200:
                return select_response

            return Response(
                {
                    "message": f"Successfully updated branch ranks for {total_updated} {year} {gender} students and selected students based on seat matrix.",
                    "selection_details": select_response.data
                },
                status=status.HTTP_200_OK
            )
        except Exception as e:
            return Response(
                {"error": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class StudentsView(APIView):
    permission_classes = [IsAuthenticated, IsRector]

    def get(self, request):
        students = User.objects.filter(user_type="student")
        serializer = UserSerializer(students, many=True)
        return Response({"data": serializer.data}, status=status.HTTP_200_OK)
class GetStudentsView(APIView):
    permission_classes = [IsAuthenticated, IsManager]

    def get(self, request):
        try:
            year = request.query_params.get('year', 'fy')
            gender = request.query_params.get('gender', 'male')
            category = request.query_params.get('category', 'all')
            branch = request.query_params.get('branch', 'all')

            valid_years = [choice[0] for choice in StudentDataEntry.CLASS_CHOICES]
            if year not in valid_years:
                return Response(
                    {"error": f"Invalid year. Must be one of: {', '.join(valid_years)}"},
                    status=status.HTTP_400_BAD_REQUEST
                )
            valid_genders = [choice[0] for choice in StudentDataEntry.GENDER_CHOICES]
            if gender not in valid_genders:
                return Response(
                    {"error": f"Invalid gender. Must be one of: {', '.join(valid_genders)}"},
                    status=status.HTTP_400_BAD_REQUEST
                )

            query = Q(class_name=year, gender=gender, verified=True)
            if category != 'all':
                query &= Q(caste__caste__iexact=category)
            if branch != 'all':
                query &= Q(branch__branch=branch)

            students = StudentDataEntry.objects.filter(query).select_related('branch', 'caste', 'admission_category').order_by('branch__branch', 'branch_rank')

            student_data = [
                {
                    "roll_no": student.roll_no,
                    "first_name": student.first_name,
                    "middle_name": student.middle_name,
                    "last_name": student.last_name,
                    "class_name": student.class_name,
                    "branch": {"name": student.branch.branch},
                    "verified": student.verified,
                    "selected": student.selected,
                    "last_selection_year": student.last_selection_year,
                    "caste": {"name": student.caste.caste},
                    "admission_category": {"name": student.admission_category.admission_category},
                }
                for student in students
            ]

            return Response({"students": student_data}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response(
                {"error": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
class SelectStudentView(APIView):
    permission_classes = [IsAuthenticated, IsManager]

    def post(self, request):
        try:
            roll_no = request.data.get('roll_no')
            if not roll_no:
                return Response(
                    {"error": "Roll number is required."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            student = StudentDataEntry.objects.get(roll_no=roll_no)
            student.selected = True
            student.is_manually_modified = True
            student.save(update_fields=['selected', 'is_manually_modified'])
            return Response(
                {"message": f"Student {roll_no} selected successfully."},
                status=status.HTTP_200_OK
            )
        except StudentDataEntry.DoesNotExist:
            return Response(
                {"error": f"Student with roll number {roll_no} not found."},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {"error": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class RemoveStudentView(APIView):
    permission_classes = [IsAuthenticated, IsManager]

    def post(self, request):
        try:
            roll_no = request.data.get('roll_no')
            if not roll_no:
                return Response(
                    {"error": "Roll number is required."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            student = StudentDataEntry.objects.get(roll_no=roll_no)
            student.selected = False
            student.is_manually_modified = True
            student.save(update_fields=['selected', 'is_manually_modified'])
            return Response(
                {"message": f"Student {roll_no} removed from selection successfully."},
                status=status.HTTP_200_OK
            )
        except StudentDataEntry.DoesNotExist:
            return Response(
                {"error": f"Student with roll number {roll_no} not found."},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {"error": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class SelectStudentsView(APIView):
    permission_classes = [IsAuthenticated, IsManager]

    def post(self, request):
        try:
            year = request.data.get('year')
            gender = request.data.get('gender', 'male')
            valid_years = [choice[0] for choice in StudentDataEntry.CLASS_CHOICES]
            if not year or year not in valid_years:
                return Response(
                    {"error": f"Invalid year. Must be one of: {', '.join(valid_years)}"},
                    status=status.HTTP_400_BAD_REQUEST
                )
            valid_genders = [choice[0] for choice in StudentDataEntry.GENDER_CHOICES]
            if gender not in valid_genders:
                return Response(
                    {"error": f"Invalid gender. Must be one of: {', '.join(valid_genders)}"},
                    status=status.HTTP_400_BAD_REQUEST
                )

            try:
                seat_matrix = SeatMatrix.objects.get(year=year, gender=gender)
            except SeatMatrix.DoesNotExist:
                return Response(
                    {"error": f"No seat matrix found for year {year} and gender {gender}"},
                    status=status.HTTP_404_NOT_FOUND
                )

            # Preserve manually modified students
            StudentDataEntry.objects.filter(
                class_name=year,
                gender=gender,
                is_manually_modified=False
            ).update(
                selected=False,
                last_selection_year=year
            )

            students = StudentDataEntry.objects.filter(
                class_name=year,
                gender=gender,
                verified=True,
                is_manually_modified=False
            ).select_related('branch', 'caste').order_by('branch_rank')

            if not students.exists():
                return Response(
                    {"message": f"No verified {year} {gender} students found to select."},
                    status=status.HTTP_200_OK
                )

            branch_seats = seat_matrix.branch_seats
            branch_groups = {}
            for student in students:
                branch_name = student.branch.branch
                if branch_name not in branch_groups:
                    branch_groups[branch_name] = []
                branch_groups[branch_name].append(student)

            total_selected = 0
            for branch_name, branch_students in branch_groups.items():
                if branch_name not in branch_seats:
                    continue

                branch_seat_info = branch_seats[branch_name]
                open_seats = branch_seat_info.get("Open", 0)
                remaining_students = branch_students.copy()
                open_selected = []
                for student in remaining_students:
                    if open_seats <= 0:
                        break
                    open_selected.append(student)
                    open_seats -= 1
                    total_selected += 1

                remaining_students = [s for s in remaining_students if s not in open_selected]
                reserved_categories = {k: v for k, v in branch_seat_info.items() if k != "Open"}
                for category, seats in reserved_categories.items():
                    category_students = [
                        s for s in remaining_students
                        if s.caste.caste.upper() == category.upper()
                    ]
                    category_students.sort(key=lambda s: s.branch_rank or float('inf'))
                    for student in category_students:
                        if seats <= 0:
                            break
                        open_selected.append(student)
                        seats -= 1
                        total_selected += 1
                    remaining_students = [s for s in remaining_students if s not in open_selected]

                for student in open_selected:
                    student.selected = True
                    student.last_selection_year = year
                    student.save(update_fields=['selected', 'last_selection_year'])

            return Response(
                {"message": f"Successfully selected {total_selected} {year} {gender} students based on seat matrix."},
                status=status.HTTP_200_OK
            )
        except Exception as e:
            return Response(
                {"error": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class exp_students(APIView):
    permission_classes = [IsAuthenticated, IsManager]

    def get(self, request):
        try:
            year = request.query_params.get('year', 'fy')
            gender = request.query_params.get('gender', 'male')
            category = request.query_params.get('category', 'all')
            export_format = request.query_params.get('exp_format', 'pdf')  # 'pdf' or 'excel'

            valid_years = [choice[0] for choice in StudentDataEntry.CLASS_CHOICES]
            if year not in valid_years:
                return Response(
                    {"error": f"Invalid year. Must be one of: {', '.join(valid_years)}"},
                    status=status.HTTP_400_BAD_REQUEST
                )
            valid_genders = [choice[0] for choice in StudentDataEntry.GENDER_CHOICES]
            if gender not in valid_genders:
                return Response(
                    {"error": f"Invalid gender. Must be one of: {', '.join(valid_genders)}"},
                    status=status.HTTP_400_BAD_REQUEST
                )

            query = Q(class_name=year, gender=gender, verified=True)
            if category != 'all':
                query &= Q(caste__caste__iexact=category)

            students = StudentDataEntry.objects.filter(query).select_related('branch', 'caste', 'admission_category').order_by('branch__branch', 'branch_rank')

            if not students.exists():
                return Response(
                    {"error": f"No verified students found for year {year}, gender {gender}, category {category}"},
                    status=status.HTTP_404_NOT_FOUND
                )

            if export_format == 'excel':
                return self.generate_excel(students, year, gender, category)

            # Default to PDF
            return self.generate_pdf(students, year, gender, category)

        except Exception as e:
            return Response(
                {"error": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def generate_pdf(self, students, year, gender, category):
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(name='Title', fontSize=16, alignment=1, spaceAfter=20)
        branch_style = ParagraphStyle(name='Branch', fontSize=14, spaceBefore=10, spaceAfter=10)

        # Title
        year_label = dict(StudentDataEntry.CLASS_CHOICES).get(year, year)
        gender_label = dict(StudentDataEntry.GENDER_CHOICES).get(gender, gender)
        category_label = category.upper() if category != 'all' else 'All Categories'
        title = Paragraph(f"Student List - {year_label} ({gender_label}, {category_label})", title_style)
        elements.append(title)

        # Group students by branch
        branch_groups = {}
        for student in students:
            branch_name = student.branch.branch
            if branch_name not in branch_groups:
                branch_groups[branch_name] = []
            branch_groups[branch_name].append(student)

        # Create a table for each branch
        for branch_name in sorted(branch_groups.keys()):
            branch_students = branch_groups[branch_name]
            elements.append(Paragraph(f"Branch: {branch_name}", branch_style))

            # Table data
            data = [['ID', 'Name', 'Category']]
            for student in branch_students:
                full_name = f"{student.first_name} {student.middle_name or ''} {student.last_name or ''}".strip()
                category = f"{student.admission_category.admission_category} ({student.caste.caste})"
                data.append([str(student.branch_rank or '-'), full_name, category])

            # Create table
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.append(table)
            elements.append(Spacer(1, 12))

        doc.build(elements)
        buffer.seek(0)
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="students_{year}_{gender}_{category}.pdf"'
        response.write(buffer.getvalue())
        buffer.close()
        return response

    def generate_excel(self, students, year, gender, category):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = f"Students_{year}_{gender}_{category}"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = openpyxl.styles.PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        alignment = Alignment(horizontal='left', vertical='center')

        # Headers
        headers = ['Branch', 'ID', 'Name', 'Category']
        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = alignment

        # Group students by branch
        branch_groups = {}
        for student in students:
            branch_name = student.branch.branch
            if branch_name not in branch_groups:
                branch_groups[branch_name] = []
            branch_groups[branch_name].append(student)

        # Write data
        row = 2
        for branch_name in sorted(branch_groups.keys()):
            for student in branch_groups[branch_name]:
                full_name = f"{student.first_name} {student.middle_name or ''} {student.last_name or ''}".strip()
                category = f"{student.admission_category.admission_category} ({student.caste.caste})"
                sheet.cell(row=row, column=1).value = branch_name
                sheet.cell(row=row, column=2).value = student.branch_rank or '-'
                sheet.cell(row=row, column=3).value = full_name
                sheet.cell(row=row, column=4).value = category
                for col in range(1, 5):
                    sheet.cell(row=row, column=col).border = border
                    sheet.cell(row=row, column=col).alignment = alignment
                row += 1

        # Adjust column widths
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            content_type='application/vnd.openxml`format`s-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="students_{year}_{gender}_{category}.xlsx"'
        response.write(buffer.getvalue())
        buffer.close()
        return response


class DashboardView(APIView):
    def get(self, request):
        try:
            year_mapping = {
                "fy": "First Year",
                "sy": "Second Year",
                "ty": "Third Year",
                "btech": "Final Year",
            }
            year_data = {
                "first-year": {"name": "First Year", "male": {}, "female": {}},
                "second-year": {"name": "Second Year", "male": {}, "female": {}},
                "third-year": {"name": "Third Year", "male": {}, "female": {}},
                "final-year": {"name": "Final Year", "male": {}, "female": {}},
            }
            overall_stats = {
                "totalSeats": 0,
                "totalRegistrations": 0,
                "totalVerified": 0,
                "totalPendingVerifications": 0,
            }
            total_registrations = StudentDataEntry.objects.count()
            total_verified = StudentDataEntry.objects.filter(verified=True).count()
            total_pending = StudentDataEntry.objects.filter(verified=None).count()
            overall_stats["totalRegistrations"] = total_registrations
            overall_stats["totalVerified"] = total_verified
            overall_stats["totalPendingVerifications"] = total_pending
            for year_key, year_name in year_mapping.items():
                class_name = year_key
                frontend_year_key = {
                    "fy": "first-year",
                    "sy": "second-year",
                    "ty": "third-year",
                    "btech": "final-year",
                }[year_key]
                for gender in ["male", "female"]:
                    seats_query = Room.objects.filter(
                        floor__class_name=class_name,
                        floor__gender=gender
                    ).aggregate(
                        room_count=Count('id'),
                        total_capacity=Sum('floor__block__per_room_capacity')
                    )
                    room_count = seats_query["room_count"] or 0
                    total_capacity = seats_query["total_capacity"] or 0
                    if room_count > 0 and total_capacity > 0:
                        per_room_capacity = total_capacity // room_count
                        total_seats = room_count * per_room_capacity
                    else:
                        total_seats = 0
                    query = StudentDataEntry.objects.filter(
                        class_name=class_name,
                        gender=gender
                    )
                    registrations = query.count()
                    verified = query.filter(verified=True).count()
                    pending_verifications = query.filter(verified=None).count()
                    if registrations > 0 and verified == 0:
                        status_value = "registration"
                    elif verified > 0 and pending_verifications > 0:
                        status_value = "verification"
                    elif verified > 0 and pending_verifications == 0 and verified < registrations:
                        status_value = "selection"
                    else:
                        status_value = "completed"
                    year_data[frontend_year_key][gender] = {
                        "totalSeats": total_seats,
                        "registrations": registrations,
                        "verified": verified,
                        "pendingVerifications": pending_verifications,
                        "status": status_value,
                    }
                male_seats = year_data[frontend_year_key]["male"]["totalSeats"]
                female_seats = year_data[frontend_year_key]["female"]["totalSeats"]
                overall_stats["totalSeats"] += male_seats + female_seats
            return Response(
                {
                    "yearData": year_data,
                    "overallStats": overall_stats,
                },
                status=status.HTTP_200_OK,
            )
        except Exception as e:
            return Response(
                {"error": f"Failed to fetch dashboard data: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )
        



from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status


import random
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.db.models import Q

class AllotRoomsView(APIView):
    def post(self, request):
        # Extract year and gender from the request body
        year = request.data.get('year')
        gender = request.data.get('gender')

        # Print the parameters to the console
        print(f"Received Year: {year}")
        print(f"Received Gender: {gender}")

        # Step 1: Convert the incoming year to the database format
        year_mapping = {
            'first': 'fy',
            'second': 'sy',
            'third': 'ty',
            'fourth': 'btech'
        }
        converted_year = year_mapping.get(year.lower())
        if not converted_year:
            return Response(
                {"error": f"Invalid year: {year}. Must be one of 'first', 'second', 'third', 'fourth'."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Validate gender
        if gender not in ['male', 'female']:
            return Response(
                {"error": f"Invalid gender: {gender}. Must be 'male' or 'female'."},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            # Step 2: Filter RoomGroup by class_name and gender, then match with StudentDataEntry
            room_groups = RoomGroup.objects.filter(class_name=converted_year, gender=gender)

            # Initialize the array to store (mis, branch_id) tuples
            current_year_toppers = []

            # Iterate through each RoomGroup
            for group in room_groups:
                mis = group.name  # The 'mis' is the name field in RoomGroup
                try:
                    student = StudentDataEntry.objects.get(roll_no=mis)
                    branch_id = student.branch.id  # Get the branch ID (ForeignKey ID)
                    current_year_toppers.append((mis, branch_id))
                except StudentDataEntry.DoesNotExist:
                    print(f"No student found with roll_no (mis): {mis}")
                    continue

            # Step 3: Get the array of branch IDs for the corresponding year and initialize special branches


            # we have to add the special branches like that 
            # we have seat allocation weight in the branch i.e it will give us values like 2.00 , 0.50, 3.00 
            #and we have to do like that if weight is 2 then two chances as we are giving right now if weight is 3 then threee consequetive chaneces if wiegt is half then 1 chance after 2 iterattions liek that 
            current_year_branches = []
            special_branches = set()  # Initialize set for special branches

            branches = Branch.objects.filter(year=converted_year)
            print(f"Found {branches.count()} branches for year {converted_year}")
            for branch in branches:
                branch_id = branch.id
                current_year_branches.append(branch_id)
                # Add to special_branches if seat_allocation_weight > 2
                if branch.seat_allocation_weight > 1:
                    special_branches.add(branch_id)


            # Print for debugging
            print("Current Year Branches:", current_year_branches)
            print("Special Branches (COMPUTER, MECHANICAL) for year", converted_year, ":", special_branches)
            print("Current Year Toppers:", current_year_toppers)

            # Step 4: Create the 3D structure (branch_id -> mis -> preferences)
            branch_mis_structure = {branch_id: [] for branch_id in current_year_branches}

            # Step 5: Populate the structure with mis and empty preferences, and collect branch ranks
            for branch_id in current_year_branches:
                mis_for_branch = [(mis, branch_id) for mis, bid in current_year_toppers if bid == branch_id]
                mis_with_ranks = []
                for mis, _ in mis_for_branch:
                    try:
                        student = StudentDataEntry.objects.get(roll_no=mis)
                        branch_rank = student.branch_rank if student.branch_rank is not None else float('inf')
                        mis_with_ranks.append((mis, branch_rank))
                    except StudentDataEntry.DoesNotExist:
                        print(f"Student with roll_no {mis} not found during ranking")
                        continue

                # Sort mis values by branch_rank (lower ranks first)
                mis_with_ranks.sort(key=lambda x: x[1])
                for mis, _ in mis_with_ranks:
                    branch_mis_structure[branch_id].append([mis, []])

            # Step 6: Replace mis with group_id and fill room preferences
            branch_group_structure = {branch_id: [] for branch_id in current_year_branches}
            for branch_id, mis_list in branch_mis_structure.items():
                for mis_entry in mis_list:
                    mis = mis_entry[0]
                    try:
                        room_group = RoomGroup.objects.get(name=mis, class_name=converted_year, gender=gender)
                        group_id = room_group.id
                        preferences = Preference.objects.filter(room_group=room_group).order_by('rank')
                        room_ids = [preference.room.id for preference in preferences]
                        branch_group_structure[branch_id].append([group_id, room_ids])
                    except RoomGroup.DoesNotExist:
                        print(f"RoomGroup with name (mis) {mis} not found")
                        continue

            # Print the resulting structure for debugging
            print("Branch Group Structure (with room preferences):", branch_group_structure)

            
            # Step 7: Allocate rooms
            # Initialize pointers and shuffle branches
            pointers = {branch_id: 0 for branch_id in current_year_branches}
            random.shuffle(current_year_branches)
            allocated_rooms = []

            # Process allocations
            while any(pointers[branch_id] < len(branch_group_structure.get(branch_id, []))
                      for branch_id in current_year_branches):
                for branch_id in current_year_branches:
                    # Determine how many groups to process (2 for special branches, 1 for others)
                    groups_to_process = 2 if branch_id in special_branches else 1
                    groups_processed = 0

                    while groups_processed < groups_to_process and pointers[branch_id] < len(branch_group_structure.get(branch_id, [])):
                        group_entry = branch_group_structure[branch_id][pointers[branch_id]]
                        group_id, room_preferences = group_entry

                        for room_id in room_preferences:
                            try:
                                room = Room.objects.get(id=room_id)
                                if not room.is_occupied:
                                    group = RoomGroup.objects.get(id=group_id)
                                    room.is_occupied = True
                                    room.alloted_group = group
                                    room.save()
                                    allocated_rooms.append({
                                        'group_id': group_id,
                                        'room_id': room_id,
                                        'room_number': room.room_id
                                    })
                                    pointers[branch_id] += 1
                                    groups_processed += 1
                                    break
                            except Room.DoesNotExist:
                                continue
            print(allocated_rooms)
            # Return the result
            return Response(
                {
                    "message": "Room allocation completed successfully",
                    "allocated_rooms": allocated_rooms,
                    "branch_group_structure": branch_group_structure
                },
                status=status.HTTP_200_OK
            )
        except Exception as e:
            print(f"Error: {str(e)}")
            return Response(
                {"error": f"An error occurred: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

#-------------------PDF GENERATION UPDATED --------------------------#
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle
from io import BytesIO
from datetime import datetime

class GeneratePDFView(APIView):
    def post(self, request):
        year = request.data.get("year")
        gender = request.data.get("gender")

        # Map frontend year values to Floor.class_name
        year_mapping = {
            "first": "fy",
            "second": "sy",
            "third": "ty",
            "fourth": "btech",
        }

        if not year or not gender:
            return Response({"error": "Year and gender are required"}, status=status.HTTP_400_BAD_REQUEST)
        
        if year not in year_mapping or gender not in ["male", "female"]:
            return Response({"error": "Invalid year or gender"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            # Convert year to class_name
            class_name = year_mapping[year]

            # Initialize dictionary to group rooms by block
            block_rooms = {}

            # Fetch all occupied rooms
            rooms = Room.objects.filter(
                is_occupied=True,
                alloted_group__isnull=False
            ).select_related("floor__block")

            # Group rooms by block for the given year and gender
            for room in rooms:
                floor = room.floor
                if floor.class_name == class_name and floor.gender == gender:
                    block_name = floor.block.name
                    if block_name not in block_rooms:
                        block_rooms[block_name] = []

                    # Fetch roommate details for this room
                    current_grp = room.alloted_group
                    roommate_details = []

                    if current_grp:
                        # Get all CustomUser objects for the group
                        roommate_users = current_grp.members.all()

                        for user in roommate_users:
                            # Check if the user has related student data
                            if hasattr(user, 'data_entry') and user.data_entry:
                                entry = user.data_entry
                                full_name = f"{entry.first_name} {entry.middle_name or ''} {entry.last_name or ''}".strip()
                                mobile_number = entry.mobile_number
                                roommate_details.append({
                                    'full_name': full_name,
                                    'mobile_number': mobile_number
                                })
                            else:
                                roommate_details.append({
                                    'full_name': user.username,  # Fallback to username if no data_entry
                                    'mobile_number': 'N/A'
                                })
                    else:
                        roommate_details.append({
                            'full_name': 'Not Allotted',
                            'mobile_number': 'N/A'
                        })

                    # Store room with its roommate details
                    block_rooms[block_name].append({
                        'room_id': room.room_id,
                        'roommates': roommate_details
                    })

            if not block_rooms:
                # print("error not")
                return Response({"error": "No rooms allotted for the selected year and gender"}, status=status.HTTP_404_NOT_FOUND)

            # Create PDF with professional layout
            buffer = BytesIO()
            p = canvas.Canvas(buffer, pagesize=A4)
            width, height = A4
            margin = 0.8 * inch
            y_position = height - margin

            # Header Section
            self._draw_header(p, width, y_position, year, gender)
            y_position -= 1.5 * inch

            # Generate tables for each block
            for block_name, rooms in block_rooms.items():
                y_position = self._draw_block_table(p, width, margin, y_position, block_name, rooms)
                
                # Check if we need a new page
                if y_position < 2 * inch:
                    p.showPage()
                    y_position = height - margin

            p.save()

            # Get PDF data
            pdf = buffer.getvalue()
            buffer.close()

            # Return PDF response
            response = HttpResponse(content_type="application/pdf")
            response["Content-Disposition"] = f'attachment; filename="COEP_Room_Allotment_{year}_{gender}_{datetime.now().year}.pdf"'
            response.write(pdf)
            return response

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def _draw_header(self, p, width, y_position, year, gender):
        """Draw the official header section"""
        # Main title
        p.setFont("Helvetica-Bold", 20)
        title = "COEP HOSTEL MANAGEMENT"
        title_width = p.stringWidth(title, "Helvetica-Bold", 20)
        p.drawString((width - title_width) / 2, y_position, title)
        
        # Subtitle
        p.setFont("Helvetica-Bold", 16)
        current_year = datetime.now().year
        subtitle = f"{year.upper()} YEAR - {gender.upper()} ROOM ALLOTMENT RESULTS"
        subtitle_width = p.stringWidth(subtitle, "Helvetica-Bold", 16)
        p.drawString((width - subtitle_width) / 2, y_position - 0.4 * inch, subtitle)
        
        # Academic year
        p.setFont("Helvetica", 12)
        academic_year = f"Academic Year: {current_year}-{current_year + 1}"
        academic_width = p.stringWidth(academic_year, "Helvetica", 12)
        p.drawString((width - academic_width) / 2, y_position - 0.7 * inch, academic_year)
        
        # Date
        date_str = f"Generated on: {datetime.now().strftime('%B %d, %Y')}"
        date_width = p.stringWidth(date_str, "Helvetica", 12)
        p.drawString((width - date_width) / 2, y_position - 0.9 * inch, date_str)
        
        # Draw a line under header
        p.line(0.8 * inch, y_position - 1.2 * inch, width - 0.8 * inch, y_position - 1.2 * inch)

    def _draw_block_table(self, p, width, margin, y_position, block_name, rooms):
        """Draw a professional table for each block"""
        # Block title
        p.setFont("Helvetica-Bold", 14)
        p.drawString(margin, y_position, f"BLOCK: {block_name}")
        y_position -= 0.4 * inch
        
        # Prepare table data
        table_data = [['Sr. No.', 'Room No.', 'Student Name', 'Mobile Number']]
        
        sr_no = 1
        for room in sorted(rooms, key=lambda x: x['room_id']):
            if room['roommates']:
                for i, roommate in enumerate(room['roommates']):
                    if i == 0:  # First roommate gets the room number
                        table_data.append([
                            str(sr_no),
                            str(room['room_id']),
                            roommate['full_name'],
                            str(roommate['mobile_number'])
                        ])
                    else:  # Additional roommates get empty room number field
                        table_data.append([
                            '',
                            '',
                            roommate['full_name'],
                            str(roommate['mobile_number'])
                        ])
                sr_no += 1
            else:
                table_data.append([
                    str(sr_no),
                    str(room['room_id']),
                    'Not Allotted',
                    'N/A'
                ])
                sr_no += 1

        # Calculate table dimensions
        table_width = width - 2 * margin
        col_widths = [0.8 * inch, 1.2 * inch, 3 * inch, 1.5 * inch]
        
        # Draw table manually for better control
        row_height = 0.3 * inch
        header_height = 0.4 * inch
        
        # Draw table borders and content
        start_x = margin
        start_y = y_position
        
        # Draw header
        p.setFont("Helvetica-Bold", 10)
        p.rect(start_x, start_y - header_height, table_width, header_height)
        
        # Header background
        p.setFillColor(colors.lightgrey)
        p.rect(start_x, start_y - header_height, table_width, header_height, fill=True)
        p.setFillColor(colors.black)
        
        # Header text
        x_pos = start_x
        for i, header in enumerate(table_data[0]):
            p.drawString(x_pos + 5, start_y - header_height + 10, header)
            x_pos += col_widths[i]
        
        # Draw vertical lines for header
        x_pos = start_x
        for width_val in col_widths:
            p.line(x_pos, start_y, x_pos, start_y - header_height)
            x_pos += width_val
        p.line(x_pos, start_y, x_pos, start_y - header_height)  # Right border
        
        current_y = start_y - header_height
        
        # Draw data rows
        p.setFont("Helvetica", 9)
        for row_idx, row in enumerate(table_data[1:]):
            # Draw row background (alternating colors)
            if row_idx % 2 == 0:
                p.setFillColor(colors.white)
            else:
                p.setFillColor(colors.Color(0.95, 0.95, 0.95))
            
            p.rect(start_x, current_y - row_height, table_width, row_height, fill=True)
            p.setFillColor(colors.black)
            
            # Draw cell content
            x_pos = start_x
            for i, cell in enumerate(row):
                # Handle text wrapping for long names
                if i == 2 and len(cell) > 25:  # Student name column
                    # Split long names
                    words = cell.split()
                    if len(words) > 2:
                        line1 = ' '.join(words[:len(words)//2])
                        line2 = ' '.join(words[len(words)//2:])
                        p.drawString(x_pos + 3, current_y - row_height + 18, line1)
                        p.drawString(x_pos + 3, current_y - row_height + 8, line2)
                    else:
                        p.drawString(x_pos + 3, current_y - row_height + 13, cell)
                else:
                    p.drawString(x_pos + 3, current_y - row_height + 13, cell)
                x_pos += col_widths[i]
            
            # Draw vertical lines
            x_pos = start_x
            for width_val in col_widths:
                p.line(x_pos, current_y, x_pos, current_y - row_height)
                x_pos += width_val
            p.line(x_pos, current_y, x_pos, current_y - row_height)  # Right border
            
            # Draw horizontal line
            p.line(start_x, current_y - row_height, start_x + table_width, current_y - row_height)
            
            current_y -= row_height
        
        return current_y - 0.5 * inch

# from django.http import HttpResponse
# from rest_framework.views import APIView
# from rest_framework.response import Response
# from rest_framework import status
# from reportlab.pdfgen import canvas
# from reportlab.lib.pagesizes import A4
# from reportlab.lib.units import inch
# from io import BytesIO

# #this is  final version 
# class GeneratePDFView(APIView):
#     def post(self, request):
#         year = request.data.get("year")
#         gender = request.data.get("gender")

#         # Map frontend year values to Floor.class_name
#         year_mapping = {
#             "first": "fy",
#             "second": "sy",
#             "third": "ty",
#             "fourth": "btech",
#         }

#         if not year or not gender:
#             return Response({"error": "Year and gender are required"}, status=status.HTTP_400_BAD_REQUEST)
        
#         if year not in year_mapping or gender not in ["male", "female"]:
#             return Response({"error": "Invalid year or gender"}, status=status.HTTP_400_BAD_REQUEST)

#         try:
#             # Convert year to class_name
#             class_name = year_mapping[year]

#             # Initialize dictionary to group rooms by block
#             block_rooms = {}

#             # Fetch all occupied rooms
#             rooms = Room.objects.filter(
#                 is_occupied=True,
#                 alloted_group__isnull=False
#             ).select_related("floor__block")

#             # Group rooms by block for the given year and gender
#             for room in rooms:
#                 floor = room.floor
#                 if floor.class_name == class_name and floor.gender == gender:
#                     block_name = floor.block.name
#                     if block_name not in block_rooms:
#                         block_rooms[block_name] = []

#                     # Fetch roommate details for this room
#                     current_grp = room.alloted_group
#                     roommate_details = []

#                     if current_grp:
#                         # Get all CustomUser objects for the group
#                         roommate_users = current_grp.members.all()

#                         for user in roommate_users:
#                             # Check if the user has related student data
#                             if hasattr(user, 'data_entry') and user.data_entry:
#                                 entry = user.data_entry
#                                 full_name = f"{entry.first_name} {entry.middle_name or ''} {entry.last_name or ''}".strip()
#                                 mobile_number = entry.mobile_number
#                                 roommate_details.append({
#                                     'full_name': full_name,
#                                     'mobile_number': mobile_number
#                                 })
#                             else:
#                                 roommate_details.append({
#                                     'full_name': user.username,  # Fallback to username if no data_entry
#                                     'mobile_number': 'N/A'
#                                 })
#                     else:
#                         roommate_details.append({
#                             'full_name': 'Not Allotted',
#                             'mobile_number': 'N/A'
#                         })

#                     # Store room with its roommate details
#                     block_rooms[block_name].append({
#                         'room_id': room.room_id,
#                         'roommates': roommate_details
#                     })

#             if not block_rooms:
#                 return Response({"error": "No rooms allotted for the selected year and gender"}, status=status.HTTP_404_NOT_FOUND)

#             # Create PDF
#             buffer = BytesIO()
#             p = canvas.Canvas(buffer, pagesize=A4)
#             width, height = A4
#             margin = inch
#             y_position = height - margin

#             # Title
#             title = f"{year.capitalize()} Year - {gender.capitalize()} Rooms"
#             p.setFont("Helvetica-Bold", 16)
#             p.drawString(margin, y_position, title)
#             y_position -= 0.5 * inch

#             # Generate sections for each block
#             for block_name, rooms in block_rooms.items():
#                 # Block heading
#                 p.setFont("Helvetica-Bold", 14)
#                 p.drawString(margin, y_position, block_name)
#                 y_position -= 0.3 * inch

#                 # List rooms and their roommates
#                 for room in sorted(rooms, key=lambda x: x['room_id']):  # Sort rooms for readability
#                     # Draw room ID
#                     p.setFont("Helvetica", 12)
#                     p.drawString(margin + 0.2 * inch, y_position, f"- {room['room_id']}")
#                     y_position -= 0.2 * inch

#                     # Draw roommate details indented below the room
#                     p.setFont("Helvetica", 10)
#                     for roommate in room['roommates']:
#                         roommate_text = f"  • {roommate['full_name']} (Mobile: {roommate['mobile_number']})"
#                         p.drawString(margin + 0.4 * inch, y_position, roommate_text)
#                         y_position -= 0.15 * inch

#                     y_position -= 0.1 * inch  # Small space after each room's roommates

#                     # Check for page overflow
#                     if y_position < margin:
#                         p.showPage()
#                         p.setFont("Helvetica", 12)  # Reset font for new page
#                         y_position = height - margin

#                 y_position -= 0.3 * inch  # Space between blocks

#                 # Check for page overflow after block
#                 if y_position < margin:
#                     p.showPage()
#                     p.setFont("Helvetica", 12)
#                     y_position = height - margin

#             p.showPage()
#             p.save()

#             # Get PDF data
#             pdf = buffer.getvalue()
#             buffer.close()

#             # Return PDF response
#             response = HttpResponse(content_type="application/pdf")
#             response["Content-Disposition"] = f'attachment; filename="room_allotment_{year}_{gender}.pdf"'
#             response.write(pdf)
#             return response

#         except Exception as e:
#             return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
# #FOR MAKING SURE OT ONLY ONE TIME ALLOTMENT IS DONE 
# from .models import AllotmentHistory
# from .serializers import AllotmentCheckSerializer, AllotmentRecordSerializer, AllotmentResetSerializer

# class CheckAllotmentView(APIView):
#     permission_classes = [IsAuthenticated]

#     def post(self, request):
#         serializer = AllotmentCheckSerializer(data=request.data)
#         if serializer.is_valid():
#             year = serializer.validated_data['year']
#             gender = serializer.validated_data['gender']
#             exists = AllotmentHistory.objects.filter(year=year, gender=gender).exists()
#             return Response({"isAllotmentDone": exists}, status=status.HTTP_200_OK)
#         return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

# class RecordAllotmentView(APIView):
#     permission_classes = [IsAuthenticated]

#     def post(self, request):
#         serializer = AllotmentRecordSerializer(data=request.data)
#         if serializer.is_valid():
#             year = serializer.validated_data['year']
#             gender = serializer.validated_data['gender']
#             AllotmentHistory.objects.get_or_create(year=year, gender=gender)
#             return Response({"message": "Allotment recorded successfully"}, status=status.HTTP_200_OK)
#         return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

# class ResetAllotmentView(APIView):
#     permission_classes = [IsAuthenticated]

#     def post(self, request):
#         serializer = AllotmentResetSerializer(data=request.data)
#         if serializer.is_valid():
#             year = serializer.validated_data['year']
#             gender = serializer.validated_data['gender']
#             AllotmentHistory.objects.filter(year=year, gender=gender).delete()
#             return Response({"message": "Allotment reset successfully"}, status=status.HTTP_200_OK)
#         return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)





# --------------- THIS IF FOR ONE TIME ALLOTMENT LOGIN-----------------------------#
from .serializers import AllotmentResetSerializer, AllotmentResetSerializer, CheckAllotmentSerializer, ManualOverrideSerializer


# New view to check if allotment has already been done
class CheckAllotmentView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        serializer = CheckAllotmentSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        year = serializer.validated_data['year']
        gender = serializer.validated_data['gender']

        year_mapping = {
            'first': 'fy',
            'second': 'sy',
            'third': 'ty',
            'fourth': 'btech'
        }
        converted_year = year_mapping.get(year.lower())
        if not converted_year:
            return Response(
                {"error": f"Invalid year: {year}. Must be one of 'first', 'second', 'third', 'fourth'."},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            allotment = AllotmentHistory.objects.get(year=converted_year, gender=gender)
            return Response({
                "is_allotted": True,
                "can_reallocate": allotment.is_manual_override,
                "allotment_date": allotment.allotment_date
            }, status=status.HTTP_200_OK)
        except AllotmentHistory.DoesNotExist:
            return Response({
                "is_allotted": False,
                "can_reallocate": False
            }, status=status.HTTP_200_OK)

# New view to record an allotment after successful allocation
class RecordAllotmentView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        serializer = AllotmentRecordSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        year = serializer.validated_data['year']
        gender = serializer.validated_data['gender']

        year_mapping = {
            'first': 'fy',
            'second': 'sy',
            'third': 'ty',
            'fourth': 'btech'
        }
        converted_year = year_mapping.get(year.lower())
        if not converted_year:
            return Response(
                {"error": f"Invalid year: {year}. Must be one of 'first', 'second', 'third', 'fourth'."},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            AllotmentHistory.objects.update_or_create(
                year=converted_year,
                gender=gender,
                defaults={'is_manual_override': False}
            )
            return Response(
                {"message": f"Allotment recorded for {year} year and {gender} gender."},
                status=status.HTTP_200_OK
            )
        except Exception as e:
            print(f"Error: {str(e)}")
            return Response(
                {"error": f"An error occurred: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

# New view to reset an allotment
class ResetAllotmentView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        serializer = AllotmentResetSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        year = serializer.validated_data['year']
        gender = serializer.validated_data['gender']

        year_mapping = {
            'first': 'fy',
            'second': 'sy',
            'third': 'ty',
            'fourth': 'btech'
        }
        converted_year = year_mapping.get(year.lower())
        if not converted_year:
            return Response(
                {"error": f"Invalid year: {year}. Must be one of 'first', 'second', 'third', 'fourth'."},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            AllotmentHistory.objects.filter(year=converted_year, gender=gender).delete()
            Room.objects.filter(
                alloted_group__class_name=converted_year,
                alloted_group__gender=gender
            ).update(is_occupied=False, alloted_group=None)

            return Response(
                {"message": f"Allotment reset successfully for {year} year and {gender} gender."},
                status=status.HTTP_200_OK
            )
        except Exception as e:
            print(f"Error: {str(e)}")
            return Response(
                {"error": f"An error occurred: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

# New view to toggle manual override
class ManualOverrideView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        serializer = ManualOverrideSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        year = serializer.validated_data['year']
        gender = serializer.validated_data['gender']
        allow_reallocation = serializer.validated_data['allow_reallocation']

        year_mapping = {
            'first': 'fy',
            'second': 'sy',
            'third': 'ty',
            'fourth': 'btech'
        }
        converted_year = year_mapping.get(year.lower())
        if not converted_year:
            return Response(
                {"error": f"Invalid year: {year}. Must be one of 'first', 'second', 'third', 'fourth'."},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            allotment = AllotmentHistory.objects.get(year=converted_year, gender=gender)
            allotment.is_manual_override = allow_reallocation
            allotment.save()
            action = "enabled" if allow_reallocation else "disabled"
            return Response(
                {"message": f"Manual re-allocation {action} for {year} year and {gender} gender."},
                status=status.HTTP_200_OK
            )
        except AllotmentHistory.DoesNotExist:
            return Response(
                {"error": f"No allotment history found for {year} year and {gender} gender."},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            print(f"Error: {str(e)}")
            return Response(
                {"error": f"An error occurred: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )